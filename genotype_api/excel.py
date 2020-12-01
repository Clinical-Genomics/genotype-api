"""Code to work with excel files"""

import logging
from pathlib import Path
from typing import ByteString, Iterable, List, Optional

import openpyxl
from genotype_api.exceptions import SexConflictError
from genotype_api.models import Analysis, Genotype
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

LOG = logging.getLogger(__name__)


class GenotypeAnalysis:
    """Class to parse and hold information about a genotype analysis

    The genotype analysis is a excel sheet with multiple individuals.
    The columns that starts with ZF predicts the gender.
    Then there are a certain number of columns with genotype information that starts with rs

    The second columns holds the sample IDs

    """

    def __init__(self, excel_file: ByteString, file_name: str, include_key: Optional[str] = None):
        LOG.info("Loading genotype information from %s", excel_file)
        self.source: str = file_name
        self.wb: Workbook = openpyxl.load_workbook(filename=excel_file)
        self.include_key: Optional[str] = include_key
        self.work_sheet: Worksheet = self.find_sheet(excel_db=self.wb)
        self.header_row: List[str] = self.get_header_cols(self.work_sheet)
        self.snp_start: int = GenotypeAnalysis.find_column(self.header_row, pattern="rs")
        self.sex_start: int = GenotypeAnalysis.find_column(self.header_row, pattern="ZF_")
        self.sex_cols: slice = slice(self.sex_start, self.sex_start + 3)
        self.rs_numbers: List[str] = self.header_row[self.snp_start :]

    @staticmethod
    def get_header_cols(sheet: Worksheet) -> List[str]:
        return [col.value for col in sheet[1]]

    @staticmethod
    def find_sheet(excel_db: Workbook, sheet_nr: int = -1) -> Worksheet:
        """Fetch the sheet nr from the database"""
        work_sheets: List[str] = excel_db.sheetnames
        sheet_name: str = work_sheets[sheet_nr]
        LOG.info("Using sheet named %s", sheet_name)
        sheet = excel_db[sheet_name]
        LOG.info("Use sheet %s", sheet)
        return sheet

    @staticmethod
    def find_column(header_row: List[str], pattern="rs") -> int:
        """Find the first column in a row that matches a pattern."""
        LOG.debug("Fetching column with %s", pattern)
        for index, column in enumerate(header_row):
            if column.startswith(pattern):
                return index

    @staticmethod
    def parse_sample_id(sample_id: str, include_key: str = None) -> Optional[str]:
        """Build samples from Excel sheet."""
        LOG.info("Parse sample id from %s using include_key: %s", sample_id, include_key)
        if include_key:
            if include_key not in sample_id:
                return None
            sample_id = sample_id.split("-")[-1]
        LOG.info("Use sample id %s", sample_id)
        return sample_id

    def generate_analyses(self) -> Iterable[Analysis]:
        """Loop over the rows and create one analysis for each individual"""
        nr_row: int
        row: List[str]
        for nr_row, row in enumerate(self.work_sheet.iter_rows()):
            if nr_row == 0:
                continue
            row_values = [cell.value for cell in row]
            ind_info = dict(zip(self.header_row, row_values))
            sample_id = GenotypeAnalysis.parse_sample_id(ind_info["SAMPLE"], self.include_key)
            if not sample_id:
                LOG.warning("Could not parse sample from row %s", nr_row)
                continue

            analysis_obj: Analysis = Analysis(
                type="genotype", source=self.source, sample_id=sample_id
            )
            analysis_obj.sex = GenotypeAnalysis.parse_sex(row_values[self.sex_cols])

            genotypes = []
            for rs_id in self.rs_numbers:
                genotypes.append(GenotypeAnalysis.build_genotype(rs_id, ind_info[rs_id]))
            analysis_obj.genotypes = genotypes

            yield analysis_obj

    @staticmethod
    def build_genotype(rs_id: str, row_value: str) -> Genotype:
        """Build genotype from Excel info."""
        alleles = row_value.split()
        return Genotype(rsnumber=rs_id, allele_1=alleles[0], allele_2=alleles[1])

    @staticmethod
    def parse_sex(sex_cells: List[str]) -> str:
        """Parse the sex prediction from a sample row."""
        predictions = set()
        # first marker
        if sex_cells[0] == "T C":
            predictions.add("male")
        elif sex_cells[0] == "C C":
            predictions.add("female")

        # second marker
        if sex_cells[1] == "T C":
            predictions.add("male")
        elif sex_cells[1] == "C C":
            predictions.add("female")

        # third marker
        if sex_cells[2] == "C T":
            predictions.add("male")
        elif sex_cells[2] == "T T":
            predictions.add("female")

        if len(predictions) == 1:
            return predictions.pop()
        elif len(predictions) == 0:
            # all assays failed
            return "unknown"
        elif len(predictions) == 2:
            # assays returned conflicting results
            message = "conflicting sex predictions: {}".format(sex_cells)
            raise SexConflictError(message)


if __name__ == "__main__":
    import sys
    import coloredlogs

    coloredlogs.install()

    in_file = Path(sys.argv[1])
    # include_key = "ID-CG-"
    include_key = "-CG-"
    genotype_analysis = GenotypeAnalysis(in_file, include_key)
    for analysis in genotype_analysis.generate_analyses():
        print(analysis)
